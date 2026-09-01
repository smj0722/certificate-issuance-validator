from __future__ import annotations
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from .models import ComparisonRow


def export_results(path: str | Path, rows: list[ComparisonRow], include_all: bool = False) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "검사결과"
    headers = ["상태", "접수번호", "오류항목", "CSI 기준", "관리프로그램", "접수대장", "사유"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E9EEF5")
        cell.alignment = Alignment(horizontal="center")
    for item in rows:
        if not include_all and item.status == "정상":
            continue
        def fmt(rec):
            if not rec:
                return "없음"
            d = rec.issue_date.isoformat() if rec.issue_date else ""
            return f"{rec.certificate_no} / {d}".strip(" /")
        ws.append([
            item.status,
            item.receipt_no,
            " · ".join(item.error_fields),
            fmt(item.csi),
            fmt(item.management),
            fmt(item.register),
            " / ".join(item.reasons),
        ])
    widths = [12, 22, 30, 34, 34, 34, 70]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64+i)].width = w
    ws.freeze_panes = "A2"
    wb.save(Path(path))

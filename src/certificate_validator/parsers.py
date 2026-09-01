from __future__ import annotations
from datetime import date, datetime
from html.parser import HTMLParser
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from .models import Record


def _date(value) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d", "%y-%m-%d"):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            pass
    return None


def _header_map(values: Iterable[object]) -> dict[str, int]:
    return {str(v).replace("\n", "").strip(): i for i, v in enumerate(values) if v not in (None, "")}


def parse_csi(path: str | Path) -> list[Record]:
    p = Path(path)
    wb = load_workbook(p, read_only=True, data_only=True)
    if "Sheet0" not in wb.sheetnames:
        raise ValueError("CSI 발급대장에 Sheet0 시트가 없습니다.")
    ws = wb["Sheet0"]
    rows = ws.iter_rows(values_only=True)
    headers = _header_map(next(rows))
    required = {"접수번호", "발급번호", "공사명", "발급일자"}
    if not required.issubset(headers):
        raise ValueError("CSI 발급대장의 필수 열을 찾을 수 없습니다.")
    company_key = "의뢰기관명" if "의뢰기관명" in headers else None
    sample_key = "봉인명" if "봉인명" in headers else None
    out: list[Record] = []
    for row in rows:
        receipt = str(row[headers["접수번호"]] or "").strip()
        if not receipt:
            continue
        out.append(Record(
            receipt_no=receipt,
            certificate_no=str(row[headers["발급번호"]] or "").strip(),
            issue_date=_date(row[headers["발급일자"]]),
            project_name=str(row[headers["공사명"]] or "").strip(),
            company_name=str(row[headers[company_key]] or "").strip() if company_key else "",
            sample_name=str(row[headers[sample_key]] or "").strip() if sample_key else "",
            source_path=p,
        ))
    return out


class _TableParser(HTMLParser):
    def __init__(self):
        super().__init__()
        self.rows: list[list[str]] = []
        self._row: list[str] | None = None
        self._cell: list[str] | None = None

    def handle_starttag(self, tag, attrs):
        if tag.lower() == "tr":
            self._row = []
        elif tag.lower() in {"td", "th"} and self._row is not None:
            self._cell = []

    def handle_data(self, data):
        if self._cell is not None:
            self._cell.append(data)

    def handle_endtag(self, tag):
        tag = tag.lower()
        if tag in {"td", "th"} and self._cell is not None and self._row is not None:
            self._row.append("".join(self._cell).strip())
            self._cell = None
        elif tag == "tr" and self._row is not None:
            if any(x.strip() for x in self._row):
                self.rows.append(self._row)
            self._row = None


def _management_rows(path: Path) -> list[list[object]]:
    raw = path.read_bytes()
    if raw.lstrip().lower().startswith(b"<html"):
        parser = _TableParser()
        text = None
        for enc in ("cp949", "euc-kr", "utf-8"):
            try:
                text = raw.decode(enc)
                break
            except UnicodeDecodeError:
                continue
        parser.feed(text or raw.decode("utf-8", errors="ignore"))
        return parser.rows
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    return [list(r) for r in ws.iter_rows(values_only=True)]


def parse_management(path: str | Path) -> tuple[list[Record], int]:
    p = Path(path)
    rows = _management_rows(p)
    header_index = None
    headers = None
    for i, row in enumerate(rows[:20]):
        hm = _header_map(row)
        if "접수번호ⓒ" in hm or "접수번호" in hm:
            header_index, headers = i, hm
            break
    if headers is None:
        raise ValueError("발행리스트의 헤더를 찾을 수 없습니다.")
    key_receipt = "접수번호ⓒ" if "접수번호ⓒ" in headers else "접수번호"
    key_cert = "성적서발급번호ⓒ" if "성적서발급번호ⓒ" in headers else "성적서발급번호"
    required = {key_receipt, key_cert, "사업자", "공사명", "시료명", "발행일"}
    if not required.issubset(headers):
        raise ValueError("발행리스트의 필수 열을 찾을 수 없습니다.")
    out: list[Record] = []
    excluded_general = 0
    for row in rows[header_index + 1:]:
        if len(row) <= max(headers.values()):
            continue
        receipt = str(row[headers[key_receipt]] or "").strip()
        if not receipt:
            continue
        if receipt.startswith("일반접수"):
            excluded_general += 1
            continue
        out.append(Record(
            receipt_no=receipt,
            certificate_no=str(row[headers[key_cert]] or "").strip(),
            issue_date=_date(row[headers["발행일"]]),
            project_name=str(row[headers["공사명"]] or "").strip(),
            company_name=str(row[headers["사업자"]] or "").strip(),
            sample_name=str(row[headers["시료명"]] or "").strip(),
            source_path=p,
        ))
    return out, excluded_general


def parse_register(path: str | Path) -> list[Record]:
    p = Path(path)
    wb = load_workbook(p, read_only=True, data_only=True, keep_vba=True)
    if "건설" not in wb.sheetnames:
        raise ValueError("접수대장에 건설 시트가 없습니다.")
    ws = wb["건설"]
    row_iter = ws.iter_rows(values_only=True)
    headers = None
    for _ in range(14):
        try:
            values = next(row_iter)
        except StopIteration:
            break
        hm = _header_map(values)
        if "접수번호" in hm and "성적서번호" in hm:
            headers = hm
            break
    if headers is None:
        raise ValueError("접수대장 건설 시트의 헤더를 찾을 수 없습니다.")
    key_revision = next((k for k in headers if k.replace(" ", "") == "수정발급일자"), None)
    out: list[Record] = []
    for row in row_iter:
        if len(row) <= max(headers.values()):
            continue
        receipt = str(row[headers["접수번호"]] or "").strip()
        if not receipt:
            continue
        out.append(Record(
            receipt_no=receipt,
            certificate_no=str(row[headers["성적서번호"]] or "").strip(),
            issue_date=_date(row[headers["발급일자"]]),
            revised_issue_date=_date(row[headers[key_revision]]) if key_revision else None,
            project_name=str(row[headers["공사명"]] or "").strip(),
            company_name=str(row[headers["시공자"]] or "").strip() if "시공자" in headers else "",
            sample_name=str(row[headers["시료명"]] or "").strip() if "시료명" in headers else "",
            source_path=p,
        ))
    return out


def classify_file(path: str | Path) -> str:
    p = Path(path)
    name = p.name
    if "품질시험 성적서 발급 대장" in name:
        parse_csi(p)
        return "csi"
    if "성적서발행발송리스트" in name:
        parse_management(p)
        return "management"
    if "접수대장" in name and p.suffix.lower() == ".xlsm":
        parse_register(p)
        return "register"
    for kind, parser in (("csi", parse_csi), ("management", lambda x: parse_management(x)[0]), ("register", parse_register)):
        try:
            parser(p)
            return kind
        except Exception:
            pass
    raise ValueError("지원하는 성적서 관련 파일로 판별할 수 없습니다.")
